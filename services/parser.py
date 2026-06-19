# services/parser.py

import openpyxl
import os
from collections import defaultdict


def safe_num(value):
    try:
        return float(value) if value is not None else 0
    except (ValueError, TypeError):
        return 0


# ─────────────────────────────────────────────
# STEP 1 — Unmerge all merged cells
# ─────────────────────────────────────────────

def unmerge_and_fill(file_path: str) -> tuple:
    wb = openpyxl.load_workbook(file_path)

    sheet_name = None
    for name in wb.sheetnames:
        if "project" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    for merge_range in list(ws.merged_cells.ranges):
        top_left_value = ws.cell(
            merge_range.min_row,
            merge_range.min_col
        ).value
        ws.unmerge_cells(str(merge_range))
        for row in ws.iter_rows(
            min_row=merge_range.min_row,
            max_row=merge_range.max_row,
            min_col=merge_range.min_col,
            max_col=merge_range.max_col,
        ):
            for cell in row:
                cell.value = top_left_value

    temp_path = file_path.replace(".xlsx", "_temp_unmerged.xlsx")
    wb.save(temp_path)
    return temp_path, sheet_name


# ─────────────────────────────────────────────
# STEP 2 — Detect header row
# ─────────────────────────────────────────────

def detect_header_row(ws) -> int:
    for row_idx in range(1, 20):
        cell_val = ws.cell(row_idx, 1).value
        if cell_val and str(cell_val).strip().lower() == "developer":
            return row_idx
    raise ValueError(
        "Could not find header row with 'Developer' in column 1."
    )


# ─────────────────────────────────────────────
# STEP 3 — Detect project columns
# ─────────────────────────────────────────────

def detect_project_layout(ws, header_row: int) -> list:
    project_layout = []
    seen_projects  = set()

    location_row      = header_row - 5
    project_row       = header_row - 4
    manager_row       = header_row - 3
    status_row        = header_row - 2
    approval_date_row = header_row - 1

    SKIP_KEYWORDS = {"ltc", "total", "others", "valdel", "grand"}

    for col in range(1, ws.max_column + 1):
        header_val = ws.cell(header_row, col).value
        if not header_val:
            continue
        if str(header_val).strip().upper() != "LTC":
            continue

        project_name  = ws.cell(project_row,       col).value
        location      = ws.cell(location_row,      col).value
        manager       = ws.cell(manager_row,        col).value
        status        = ws.cell(status_row,         col).value
        approval_date = ws.cell(approval_date_row,  col).value

        if not project_name:
            continue

        project_name_clean = str(project_name).strip()

        if project_name_clean in seen_projects:
            continue
        if project_name_clean.lower() in SKIP_KEYWORDS:
            continue

        location_clean = str(location).strip().lower() if location else ""
        if location_clean == project_name_clean.lower():
            continue

        seen_projects.add(project_name_clean)

        next_header = str(
            ws.cell(header_row, col + 1).value or ""
        ).strip().upper()

        has_others = next_header not in ("LTC", "TOTAL", "")

        project_layout.append({
            "name":          project_name_clean,
            "location":      str(location).strip() if location else "",
            "manager":       str(manager).strip()  if manager  else "",
            "status":        str(status).strip()   if status   else "",
            "approval_date": approval_date,
            "ltc_col":       col,
            "valdel_col":    col + 1 if has_others else None,
            "total_p_col":   col + 2 if has_others else None,
        })

    return project_layout


# ─────────────────────────────────────────────
# STEP 4 — Detect summary columns
# ─────────────────────────────────────────────

def detect_summary_columns(ws, header_row: int) -> dict:
    summary = {}

    for col in range(1, ws.max_column + 1):
        cell_val = ws.cell(header_row, col).value
        if not cell_val:
            continue
        val_clean = str(cell_val).strip().lower()

        if "own" in val_clean:
            summary["own_lic_col"] = col
        elif "lease" in val_clean:
            summary["lease_lic_col"] = col
        elif val_clean == "total":
            summary["total_col"] = col
        elif "loc" in val_clean and "1" in val_clean and "ltc" in val_clean:
            summary["loc1_ltc_col"] = col
        elif "loc" in val_clean and "2" in val_clean and "ltc" in val_clean:
            summary["loc2_ltc_col"] = col

    return summary


# ─────────────────────────────────────────────
# STEP 5 — Parse all data rows
# ─────────────────────────────────────────────

def parse_source(ws, header_row, project_layout, summary_cols) -> list:
    records     = []
    current_dev = None
    current_sw  = None

    for row_idx in range(header_row + 1, ws.max_row + 1):
        dev  = ws.cell(row_idx, 1).value
        sw   = ws.cell(row_idx, 2).value
        dept = ws.cell(row_idx, 3).value

        # Read own/lease/total from total row then skip it
        if dept and str(dept).strip().lower() == "total":
            own_lic_val   = safe_num(ws.cell(row_idx, summary_cols.get("own_lic_col",   0)).value) if summary_cols.get("own_lic_col")   else 0
            lease_lic_val = safe_num(ws.cell(row_idx, summary_cols.get("lease_lic_col", 0)).value) if summary_cols.get("lease_lic_col") else 0
            total_lic_val = safe_num(ws.cell(row_idx, summary_cols.get("total_col",     0)).value) if summary_cols.get("total_col")     else 0
            for r in records:
                if r["software"] == current_sw:
                    r["own_lic"]   = own_lic_val
                    r["lease_lic"] = lease_lic_val
                    r["total_lic"] = total_lic_val
            continue

        # Skip fully empty rows
        if dev is None and sw is None and dept is None:
            continue

        # ffill developer and software
        if dev: current_dev = str(dev).strip()
        if sw:  current_sw  = str(sw).strip()

        if dept is None:
            continue

        # Read per-project licence values
        proj_data = {}
        grand_ltc = 0

        for p in project_layout:
            ltc_v = safe_num(ws.cell(row_idx, p["ltc_col"]).value)

            valdel_v = 0
            if p["valdel_col"] is not None:
                valdel_v = safe_num(ws.cell(row_idx, p["valdel_col"]).value)

            total_p = 0
            if p["total_p_col"] is not None:
                total_p = safe_num(ws.cell(row_idx, p["total_p_col"]).value)

            grand_ltc += ltc_v  # always accumulate LTC

            proj_data[p["name"]] = {
                "ltc":     ltc_v,
                "valdel":  valdel_v,
                "total_p": total_p,
            }

        others = safe_num(ws.cell(row_idx, summary_cols.get("loc1_ltc_col", 0)).value) if summary_cols.get("loc1_ltc_col") else 0

        records.append({
            "developer": current_dev,
            "software":  current_sw,
            "dept":      str(dept).strip(),
            "projects":  proj_data,
            "grand_ltc": grand_ltc,
            "others":    others,
            "total_lic": 0,   # filled in when total row is read
            "own_lic":   0,   # filled in when total row is read
            "lease_lic": 0,   # filled in when total row is read
        })

    return records


# ─────────────────────────────────────────────
# STEP 6 — Aggregate by software
# ─────────────────────────────────────────────

def aggregate_by_software(records: list) -> dict:
    agg          = {}
    seen_sw      = set()

    for r in records:
        sw = r["software"]
        if not sw:
            continue

        if sw not in agg:
            agg[sw] = {
                "developer": "",
                "total_lic": 0,
                "lease_lic": 0,
                "own_lic":   0,
                "depts":     [],
            }

        agg[sw]["developer"] = r["developer"]
        agg[sw]["depts"].append(r)

        # Set own/lease/total only once per software
        if sw not in seen_sw:
            agg[sw]["own_lic"]   = r["own_lic"]
            agg[sw]["lease_lic"] = r["lease_lic"]
            agg[sw]["total_lic"] = r["total_lic"]
            seen_sw.add(sw)

    return agg


# ─────────────────────────────────────────────
# MAIN ENTRY POINT
# ─────────────────────────────────────────────

def parse_excel(file_path: str) -> tuple:
    from core.errors import InvalidFileError, SheetNotFoundError

    temp_path = None

    try:
        temp_path, sheet_name = unmerge_and_fill(file_path)
        wb         = openpyxl.load_workbook(temp_path)
        ws         = wb[sheet_name]
        header_row = detect_header_row(ws)

        project_layout = detect_project_layout(ws, header_row)
        if not project_layout:
            raise InvalidFileError("No project columns found.")

        summary_cols = detect_summary_columns(ws, header_row)
        records      = parse_source(ws, header_row, project_layout, summary_cols)

        if not records:
            raise InvalidFileError("No data rows found.")

        sw_agg = aggregate_by_software(records)
        return project_layout, records, sw_agg

    except (InvalidFileError, SheetNotFoundError):
        raise
    except ValueError as e:
        raise SheetNotFoundError(str(e))
    except Exception as e:
        raise InvalidFileError(f"Unexpected error parsing file: {str(e)}")
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)