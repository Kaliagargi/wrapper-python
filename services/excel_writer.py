# services/excel_writer.py

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
TITLE_FONT  = Font(bold=True, name="Arial", size=12, color="1F4E79")
TOTAL_FILL  = PatternFill("solid", fgColor="FFF2CC")
TOTAL_FONT  = Font(bold=True, name="Arial", size=10)
ALT_FILL    = PatternFill("solid", fgColor="D9E2F3")
WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
RED_FONT    = Font(bold=True, name="Arial", size=10, color="C00000")
GREEN_FONT  = Font(bold=True, name="Arial", size=10, color="375623")
NORMAL_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

OUTPUT_DIR = "outputs"
os.makedirs(OUTPUT_DIR, exist_ok=True)


# ─────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────

def hdr(ws, row, col, value):
    c = ws.cell(row, col, value)
    c.font      = HEADER_FONT
    c.fill      = HEADER_FILL
    c.border    = THIN_BORDER
    c.alignment = Alignment(horizontal="center", vertical="center")
    return c


def val(ws, row, col, value, alt=False, left=False):
    c = ws.cell(row, col, value)
    c.font      = NORMAL_FONT
    c.fill      = ALT_FILL if alt else WHITE_FILL
    c.border    = THIN_BORDER
    c.alignment = Alignment(
        horizontal="left" if left else "center",
        vertical="center"
    )
    return c


def title_row(ws, row, text, ncols):
    c = ws.cell(row, 1, text)
    c.font      = TITLE_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=ncols
    )
    ws.row_dimensions[row].height = 22


def total_row(ws, row, label, values: dict, ncols):
    c = ws.cell(row, 1, label)
    c.font      = TOTAL_FONT
    c.fill      = TOTAL_FILL
    c.border    = THIN_BORDER
    c.alignment = Alignment(horizontal="left", vertical="center")
    for col, v in values.items():
        c2 = ws.cell(row, col, v)
        c2.font      = TOTAL_FONT
        c2.fill      = TOTAL_FILL
        c2.border    = THIN_BORDER
        c2.alignment = Alignment(horizontal="center", vertical="center")


def gap_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row, col, "")
        c.fill = WHITE_FILL
    ws.row_dimensions[row].height = 8


def get_or_create_sheet(wb, sheet_name):
    """Delete sheet if exists, create fresh."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    return wb.create_sheet(sheet_name)


# ─────────────────────────────────────────────
# SHEET 1 — Licence Summary
# ─────────────────────────────────────────────

def write_sheet1(wb, table1_data: list):
    ws    = get_or_create_sheet(wb, "Licence Summary")
    ncols = 8          # Developer Software Total Own Lease Annual Advent Order
    row   = 1

    title_row(ws, row, "📊  Licence Summary", ncols)
    row += 1

    for ci, h in enumerate(
        ["Developer", "Software", "Total Lic", "Own Lic",
         "Lease Lic", "Annual", "Advent", "Order"], 1
    ):
        hdr(ws, row, ci, h)
    row += 1

    data_start = row
    for ri, d in enumerate(table1_data):
        alt = ri % 2 == 0
        val(ws, row, 1, d["developer"],           alt=alt, left=True)
        val(ws, row, 2, d["software"],             alt=alt, left=True)
        val(ws, row, 3, d["total_lic"],            alt=alt)
        val(ws, row, 4, d["own_lic"],              alt=alt)
        val(ws, row, 5, d["lease_lic"],            alt=alt)
        val(ws, row, 6, d["annual"],               alt=alt)
        val(ws, row, 7, d.get("advent", 0),        alt=alt)
        val(ws, row, 8, d["order_lic"],            alt=alt)
        row += 1

    total_row(ws, row, "TOTAL", {
        ci: f"=SUM({get_column_letter(ci)}{data_start}:{get_column_letter(ci)}{row-1})"
        for ci in range(3, 9)
    }, ncols)

    for ci, w in zip("ABCDEFGH", [16, 20, 12, 12, 12, 12, 12, 12]):
        ws.column_dimensions[ci].width = w
    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────
# SHEET 2 — Allocated
# ─────────────────────────────────────────────

def write_sheet2(wb, table2_data: dict):
    ws    = get_or_create_sheet(wb, "Allocated")
    ncols = 3
    row   = 1

    title_row(ws, row, "📋  Allocated Licences", ncols)
    row += 1

    for sw, data in table2_data.items():
        c = ws.cell(row, 1, f"{data['developer']} — {sw}")
        c.font      = Font(bold=True, name="Arial", size=11, color="2E75B6")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row,   end_column=ncols
        )
        row += 1

        for ci, h in enumerate(["Category", "Description", "Value"], 1):
            hdr(ws, row, ci, h)
        row += 1

        for ri, r in enumerate(data["rows"]):
            alt      = ri % 2 == 0
            is_total = r["label"] == "Total"

            if is_total:
                total_row(ws, row, r["label"], {
                    2: r.get("description", ""),
                    3: r["value"]
                }, ncols)
            else:
                val(ws, row, 1, r["label"],                  alt=alt, left=True)
                val(ws, row, 2, r.get("description", ""),    alt=alt, left=True)
                val(ws, row, 3, r["value"],                  alt=alt)
            row += 1

        gap_row(ws, row, ncols)
        row += 2

    for ci, w in zip("ABC", [20, 30, 14]):
        ws.column_dimensions[ci].width = w
    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────
# SHEET 3 — Required
# ─────────────────────────────────────────────

def write_sheet3(wb, table3_data: dict):
    ws    = get_or_create_sheet(wb, "Required")
    ncols = 3
    row   = 1

    title_row(ws, row, "🔢  Required Licences", ncols)
    row += 1

    for sw, data in table3_data.items():
        c = ws.cell(row, 1, f"{data['developer']} — {sw}  |  Order: {data['order']}")
        c.font      = Font(bold=True, name="Arial", size=11, color="2E75B6")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row,   end_column=ncols
        )
        row += 1

        for ci, h in enumerate(["Category", "Value", ""], 1):
            hdr(ws, row, ci, h)
        row += 1

        for ri, r in enumerate(data["rows"]):
            alt      = ri % 2 == 0
            is_total = r["label"] == "Total"
            is_vec   = r["label"] == "VEC"

            if is_total:
                total_row(ws, row, r["label"], {2: r["value"]}, ncols)
            else:
                val(ws, row, 1, r["label"], alt=alt, left=True)
                c = ws.cell(row, 2, r["value"])
                c.fill      = ALT_FILL if alt else WHITE_FILL
                c.border    = THIN_BORDER
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.font      = RED_FONT if (is_vec and r["value"] < 0) else (
                              GREEN_FONT if is_vec else NORMAL_FONT)
                val(ws, row, 3, "", alt=alt)
            row += 1

        gap_row(ws, row, ncols)
        row += 2

    for ci, w in zip("ABC", [20, 14, 8]):
        ws.column_dimensions[ci].width = w
    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────
# SHEET 4 — ISL
# ─────────────────────────────────────────────

def write_sheet4(wb, table4_data: dict):
    ws    = get_or_create_sheet(wb, "ISL")
    ncols = 4
    row   = 1

    title_row(ws, row, "🏦  ISL — In Stock Licences", ncols)
    row += 1

    for sw, data in table4_data.items():
        c = ws.cell(row, 1, f"{data['developer']} — {sw}")
        c.font      = Font(bold=True, name="Arial", size=11, color="2E75B6")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(
            start_row=row, start_column=1,
            end_row=row,   end_column=ncols
        )
        row += 1

        for ci, h in enumerate(["Dept", "LTM", "Share", "Total"], 1):
            hdr(ws, row, ci, h)
        row += 1

        for ri, r in enumerate(data["rows"]):
            alt      = ri % 2 == 0
            is_total = r["dept"] == "TOTAL"

            if is_total:
                total_row(ws, row, "TOTAL", {
                    2: r["ltm"],
                    3: r["share"],
                    4: r["total"],
                }, ncols)
            else:
                val(ws, row, 1, r["dept"],  alt=alt, left=True)
                val(ws, row, 2, r["ltm"],   alt=alt)
                val(ws, row, 3, r["share"], alt=alt)
                val(ws, row, 4, r["total"], alt=alt)
            row += 1

        gap_row(ws, row, ncols)
        row += 2

    for ci, w in zip("ABCD", [16, 12, 12, 12]):
        ws.column_dimensions[ci].width = w
    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────
# SHEET 5 — Keystore (active keys only)
# ─────────────────────────────────────────────

def write_sheet5(wb, keystore_data: dict):
    ws    = get_or_create_sheet(wb, "Keystore")
    ncols = 5
    row   = 1

    title_row(ws, row, "🔑  Keystore", ncols)
    row += 1

    for ci, h in enumerate(
        ["Developer", "Software", "Label", "Key ID", "Value"], 1
    ):
        hdr(ws, row, ci, h)
    row += 1

    data_start = row

    for sw, data in keystore_data.items():
        # Only write active keys
        active_keys = [k for k in data["keys"] if k["active"]]

        for ri, k in enumerate(active_keys):
            alt = ri % 2 == 0
            val(ws, row, 1, data["developer"], alt=alt, left=True)
            val(ws, row, 2, sw,                alt=alt, left=True)
            val(ws, row, 3, k["label"],        alt=alt, left=True)
            val(ws, row, 4, k["key_id"],       alt=alt, left=True)
            val(ws, row, 5, k["value"] if k["value"] is not None else "User Input", alt=alt)
            row += 1

        gap_row(ws, row, ncols)
        row += 2

    for ci, w in zip("ABCDE", [16, 20, 16, 16, 14]):
        ws.column_dimensions[ci].width = w
    ws.freeze_panes = "A3"


# ─────────────────────────────────────────────
# GET OUTPUT PATH
# ─────────────────────────────────────────────

def get_output_path(file_path: str) -> str:
    """Returns the output Excel path based on input filename."""
    base_name   = os.path.basename(file_path)
    output_name = base_name.replace(".xlsx", "_report.xlsx")
    return os.path.join(OUTPUT_DIR, output_name)


# ─────────────────────────────────────────────
# SAVE SHEET — called every time user builds a table
# ─────────────────────────────────────────────

def save_sheet(
    file_path:  str,
    sheet_name: str,
    data,
    extra=None,
):
    """
    Updates a single sheet in the report Excel file.
    Creates the file if it doesn't exist yet.
    Called every time user builds any table.

    sheet_name options:
      "Licence Summary" → data is table1_data (list)
      "Allocated"       → data is table2_data (dict)
      "Required"        → data is table3_data (dict)
      "ISL"             → data is table4_data (dict)
      "Keystore"        → data is keystore_data (dict)
    """
    output_path = get_output_path(file_path)

    # Load existing report or create from input file
    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
    else:
        wb = openpyxl.load_workbook(file_path)

    if sheet_name == "Licence Summary":
        write_sheet1(wb, data)
    elif sheet_name == "Allocated":
        write_sheet2(wb, data)
    elif sheet_name == "Required":
        write_sheet3(wb, data)
    elif sheet_name == "ISL":
        write_sheet4(wb, data)
    elif sheet_name == "Keystore":
        write_sheet5(wb, data)

    wb.save(output_path)
    return output_path


# ─────────────────────────────────────────────
# GENERATE FULL REPORT — called on download
# ─────────────────────────────────────────────

def generate_report(
    file_path:    str,
    table1_data:  list,
    table2_data:  dict,
    table3_data:  dict,
    table4_data:  dict,
    keystore_data: dict,
) -> str:
    """
    Writes all 5 sheets to the report Excel file.
    Called when user clicks Download.
    Just pushes the file — no recomputation.
    """
    output_path = get_output_path(file_path)

    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
    else:
        wb = openpyxl.load_workbook(file_path)

    write_sheet1(wb, table1_data)
    write_sheet2(wb, table2_data)
    write_sheet3(wb, table3_data)
    write_sheet4(wb, table4_data)
    write_sheet5(wb, keystore_data)

    wb.save(output_path)
    return output_path